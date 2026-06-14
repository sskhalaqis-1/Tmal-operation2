#!/usr/bin/env python3
from __future__ import annotations

import csv
import io
import json
import mimetypes
import os
import re
import secrets
import sys
import uuid
from datetime import datetime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, urlparse

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - startup message handles this
    load_workbook = None


ROOT = Path(__file__).resolve().parent
APP_DIR = ROOT 
DEPLOY_DATA_DIR = os.environ.get("COFFEE_OPS_DATA_DIR") or os.environ.get("RAILWAY_VOLUME_MOUNT_PATH")
DATA_DIR = Path(DEPLOY_DATA_DIR).resolve() if DEPLOY_DATA_DIR else ROOT / "data"
STORE_FILE = DATA_DIR / "store.json"
CONFIG_FILE = DATA_DIR / "config.json"
UPLOAD_DIR = DATA_DIR / "uploads"

USERS = {
    "admin": {"role": "admin", "password": os.environ.get("COFFEE_OPS_ADMIN_PASSWORD", "admin123")},
    "staff": {"role": "staff", "pin": os.environ.get("COFFEE_OPS_STAFF_PIN", "2468")},
}
TOKENS: dict[str, dict] = {}


def now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def read_json(path: Path, fallback):
    if not path.exists():
        return fallback
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def write_json(path: Path, data) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp = path.with_suffix(".tmp")
    with temp.open("w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)
    temp.replace(path)


def default_store() -> dict:
    return {
        "checklists": [],
        "inventory_counts": [],
        "issues": [],
        "sales_uploads": [],
    }


def load_store() -> dict:
    store = read_json(STORE_FILE, default_store())
    for key, value in default_store().items():
        store.setdefault(key, value)
    return store


def ensure_runtime_files() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    if DEPLOY_DATA_DIR and not CONFIG_FILE.exists():
        source_config = ROOT / "data" / "config.json"
        if source_config.exists():
            write_json(CONFIG_FILE, read_json(source_config, {}))
    if not STORE_FILE.exists():
        save_store(default_store())


def save_store(store: dict) -> None:
    write_json(STORE_FILE, store)


def make_id(prefix: str) -> str:
    return f"{prefix}_{uuid.uuid4().hex[:10]}"


def send_json(handler: BaseHTTPRequestHandler, data, status: int = 200) -> None:
    body = json.dumps(data).encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", "application/json")
    handler.send_header("Content-Length", str(len(body)))
    handler.end_headers()
    handler.wfile.write(body)


def parse_body(handler: BaseHTTPRequestHandler) -> dict:
    length = int(handler.headers.get("Content-Length", "0"))
    raw = handler.rfile.read(length)
    if not raw:
        return {}
    return json.loads(raw.decode("utf-8"))


def request_user(handler: BaseHTTPRequestHandler) -> dict | None:
    auth = handler.headers.get("Authorization", "")
    if not auth.startswith("Bearer "):
        return None
    return TOKENS.get(auth.removeprefix("Bearer ").strip())


def require_role(handler: BaseHTTPRequestHandler, allowed_roles: set[str]) -> dict | None:
    user = request_user(handler)
    if not user or user.get("role") not in allowed_roles:
        send_json(handler, {"error": "Unauthorized"}, 401)
        return None
    return user


def multipart_file(handler: BaseHTTPRequestHandler) -> tuple[str, bytes]:
    content_type = handler.headers.get("Content-Type", "")
    match = re.search(r"boundary=([^;]+)", content_type)
    if not match:
        raise ValueError("Missing multipart boundary")
    boundary = match.group(1).strip('"').encode()
    raw = handler.rfile.read(int(handler.headers.get("Content-Length", "0")))
    for part in raw.split(b"--" + boundary):
        if b"filename=" not in part:
            continue
        header, _, content = part.partition(b"\r\n\r\n")
        name_match = re.search(rb'filename="([^"]+)"', header)
        filename = name_match.group(1).decode("utf-8", errors="replace") if name_match else "upload"
        return filename, content.rstrip(b"\r\n-")
    raise ValueError("No file found in upload")


def normalize_header(value) -> str:
    return str(value or "").strip().lower().replace("_", " ")


def parse_sales_rows(filename: str, content: bytes) -> list[dict]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        text = content.decode("utf-8-sig")
        return list(csv.DictReader(io.StringIO(text)))
    if suffix in {".xlsx", ".xlsm"}:
        if load_workbook is None:
            raise ValueError("openpyxl is not available in this Python environment")
        workbook = load_workbook(io.BytesIO(content), data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(cell or "").strip() for cell in rows[0]]
        result = []
        for row in rows[1:]:
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue
            result.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers))})
        return result
    raise ValueError("Upload a .xlsx, .xlsm, or .csv file")


def number_from(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = re.sub(r"[^0-9.\-]", "", str(value))
    try:
        return float(cleaned) if cleaned else 0.0
    except ValueError:
        return 0.0


def summarize_sales(filename: str, content: bytes) -> dict:
    rows = parse_sales_rows(filename, content)
    if not rows:
        return {"total_sales": 0, "order_count": 0, "top_products": [], "sales_by_category": [], "detected_columns": []}

    headers = list(rows[0].keys())
    header_map = {normalize_header(h): h for h in headers}

    def pick(*names):
        for name in names:
            if name in header_map:
                return header_map[name]
        for normalized, original in header_map.items():
            if any(name in normalized for name in names):
                return original
        return None

    product_col = pick("product", "item", "item name", "name")
    category_col = pick("category", "department", "group")
    order_col = pick("order id", "order", "receipt", "ticket", "invoice")
    qty_col = pick("quantity", "qty", "count")
    sales_col = pick("total", "sales", "amount", "net sales", "gross sales", "price")

    total_sales = sum(number_from(row.get(sales_col)) for row in rows) if sales_col else 0
    order_values = {str(row.get(order_col)).strip() for row in rows if row.get(order_col)} if order_col else set()
    order_count = len(order_values) if order_values else len(rows)

    products: dict[str, dict] = {}
    categories: dict[str, float] = {}
    for row in rows:
        product = str(row.get(product_col) or "Unknown product").strip()
        category = str(row.get(category_col) or "Uncategorized").strip()
        qty = number_from(row.get(qty_col)) if qty_col else 1
        sales = number_from(row.get(sales_col)) if sales_col else 0
        products.setdefault(product, {"name": product, "quantity": 0, "sales": 0})
        products[product]["quantity"] += qty
        products[product]["sales"] += sales
        categories[category] = categories.get(category, 0) + sales

    return {
        "total_sales": round(total_sales, 2),
        "order_count": order_count,
        "top_products": sorted(products.values(), key=lambda item: (item["quantity"], item["sales"]), reverse=True)[:8],
        "sales_by_category": [{"category": k, "sales": round(v, 2)} for k, v in sorted(categories.items(), key=lambda item: item[1], reverse=True)],
        "detected_columns": headers,
    }


def inventory_comparison(store: dict, current: dict) -> list[dict]:
    previous = None
    for count in reversed(store["inventory_counts"]):
        if count.get("branch") == current.get("branch"):
            previous = count
            break

    previous_items = {item["id"]: item for item in previous.get("items", [])} if previous else {}
    comparison = []
    for item in current.get("items", []):
        before = previous_items.get(item["id"], {}).get("quantity")
        delta = None if before is None else item["quantity"] - before
        unusual = bool(delta is not None and abs(delta) >= max(10, abs(before) * 0.35))
        comparison.append({**item, "previous": before, "delta": delta, "unusual": unusual})
    return comparison


def dashboard(store: dict) -> dict:
    today = datetime.now().date().isoformat()
    today_checklists = [item for item in store["checklists"] if item.get("created_at", "").startswith(today)]
    today_inventory = [item for item in store["inventory_counts"] if item.get("created_at", "").startswith(today)]
    open_issues = [item for item in store["issues"] if item.get("status") != "Resolved"]
    latest_sales = store["sales_uploads"][-1] if store["sales_uploads"] else None
    weekly_sales = sum(upload.get("summary", {}).get("total_sales", 0) for upload in store["sales_uploads"][-7:])
    return {
        "today": today,
        "checklists_today": today_checklists,
        "inventory_today": today_inventory,
        "open_issues": open_issues,
        "latest_sales": latest_sales,
        "weekly_summary": {
            "checklists": len([x for x in store["checklists"] if x.get("created_at", "")[:10] >= today[:8] + "01"]),
            "inventory_counts": len(store["inventory_counts"][-7:]),
            "open_issues": len(open_issues),
            "sales_total_recent_uploads": round(weekly_sales, 2),
        },
    }


class CoffeeOpsHandler(BaseHTTPRequestHandler):
    def log_message(self, fmt, *args):
        print(f"{self.address_string()} - {fmt % args}")

    def do_HEAD(self):
        parsed = urlparse(self.path)
        path = APP_DIR / ("index.html" if parsed.path == "/" else parsed.path.lstrip("/"))
        if not path.exists() or not path.is_file():
            self.send_error(404)
            return
        mime, _ = mimetypes.guess_type(str(path))
        self.send_response(200)
        self.send_header("Content-Type", mime or "application/octet-stream")
        self.send_header("Content-Length", str(path.stat().st_size))
        self.end_headers()

    def do_GET(self):
        parsed = urlparse(self.path)
        if parsed.path == "/health":
            send_json(self, {"ok": True})
            return
        if parsed.path == "/api/config":
            send_json(self, read_json(CONFIG_FILE, {}))
            return
        if parsed.path == "/api/state":
            if not require_role(self, {"admin"}):
                return
            send_json(self, load_store())
            return
        if parsed.path == "/api/dashboard":
            if not require_role(self, {"admin"}):
                return
            send_json(self, dashboard(load_store()))
            return
        self.serve_static(parsed.path)

    def do_POST(self):
        parsed = urlparse(self.path)
        store = load_store()
        try:
            if parsed.path == "/api/login":
                data = parse_body(self)
                role = data.get("role")
                employee = str(data.get("employee") or "").strip()
                if role == "admin" and data.get("password") == USERS["admin"]["password"]:
                    token = secrets.token_urlsafe(24)
                    TOKENS[token] = {"role": "admin", "employee": employee or "Admin"}
                    send_json(self, {"token": token, "role": "admin", "employee": employee or "Admin"})
                    return
                if role == "staff" and data.get("pin") == USERS["staff"]["pin"] and employee:
                    token = secrets.token_urlsafe(24)
                    TOKENS[token] = {"role": "staff", "employee": employee}
                    send_json(self, {"token": token, "role": "staff", "employee": employee})
                    return
                send_json(self, {"error": "Invalid login"}, 401)
                return
            if parsed.path == "/api/logout":
                auth = self.headers.get("Authorization", "")
                if auth.startswith("Bearer "):
                    TOKENS.pop(auth.removeprefix("Bearer ").strip(), None)
                send_json(self, {"ok": True})
                return
            if parsed.path == "/api/checklists":
                user = require_role(self, {"admin", "staff"})
                if not user:
                    return
                data = parse_body(self)
                data["employee"] = data.get("employee") or user.get("employee")
                data.update({"id": make_id("check"), "created_at": now_iso()})
                store["checklists"].append(data)
                save_store(store)
                send_json(self, data, 201)
                return
            if parsed.path == "/api/inventory":
                user = require_role(self, {"admin", "staff"})
                if not user:
                    return
                data = parse_body(self)
                data["employee"] = data.get("employee") or user.get("employee")
                data.update({"id": make_id("inv"), "created_at": now_iso()})
                data["comparison"] = inventory_comparison(store, data)
                store["inventory_counts"].append(data)
                save_store(store)
                send_json(self, data, 201)
                return
            if parsed.path == "/api/issues":
                user = require_role(self, {"admin", "staff"})
                if not user:
                    return
                data = parse_body(self)
                data["employee"] = data.get("employee") or user.get("employee")
                data.update({"id": make_id("issue"), "status": "Open", "comments": [], "created_at": now_iso()})
                store["issues"].append(data)
                save_store(store)
                send_json(self, data, 201)
                return
            if parsed.path == "/api/sales-upload":
                if not require_role(self, {"admin"}):
                    return
                filename, content = multipart_file(self)
                UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
                saved_name = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{Path(filename).name}"
                (UPLOAD_DIR / saved_name).write_bytes(content)
                upload = {
                    "id": make_id("sales"),
                    "filename": filename,
                    "stored_file": saved_name,
                    "created_at": now_iso(),
                    "summary": summarize_sales(filename, content),
                }
                store["sales_uploads"].append(upload)
                save_store(store)
                send_json(self, upload, 201)
                return
        except Exception as exc:
            send_json(self, {"error": str(exc)}, 400)
            return
        send_json(self, {"error": "Not found"}, 404)

    def do_PATCH(self):
        parsed = urlparse(self.path)
        match = re.match(r"^/api/issues/([^/]+)$", parsed.path)
        if not match:
            send_json(self, {"error": "Not found"}, 404)
            return
        data = parse_body(self)
        if not require_role(self, {"admin"}):
            return
        store = load_store()
        for issue in store["issues"]:
            if issue["id"] == match.group(1):
                if "status" in data:
                    issue["status"] = data["status"]
                if data.get("comment"):
                    issue.setdefault("comments", []).append({"text": data["comment"], "created_at": now_iso()})
                issue["updated_at"] = now_iso()
                save_store(store)
                send_json(self, issue)
                return
        send_json(self, {"error": "Issue not found"}, 404)

    def serve_static(self, request_path: str):
        path = APP_DIR / ("index.html" if request_path == "/" else request_path.lstrip("/"))
        if not path.exists() or not path.is_file():
            self.send_error(404)
            return
        content = path.read_bytes()
        mime, _ = mimetypes.guess_type(str(path))
        self.send_response(200)
        self.send_header("Content-Type", mime or "application/octet-stream")
        self.send_header("Content-Length", str(len(content)))
        self.end_headers()
        self.wfile.write(content)


if __name__ == "__main__":
    ensure_runtime_files()
    port = int(os.environ.get("PORT", "8000"))
    host = os.environ.get("HOST", "0.0.0.0")
    server = ThreadingHTTPServer((host, port), CoffeeOpsHandler)
    shown_host = "127.0.0.1" if host == "0.0.0.0" else host
    print(f"Coffee Ops MVP running at http://{shown_host}:{port}")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nStopped")
        sys.exit(0)
